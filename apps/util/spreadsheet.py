"""Spreadsheet class is used to make Physics and Astronomy Marksheets."""

# Python imports
import os
import re
from abc import abstractmethod
from tempfile import NamedTemporaryFile
from textwrap import shorten

# Django imports
from django.db.models import Count
from django.http import HttpResponse

# external imports
import numpy as np
import openpyxl as opx
from constance import config
from openpyxl.styles import PatternFill

# Set up some fill colours

red_fill = PatternFill(patternType="solid", start_color="FF0000")
green_fill = PatternFill(patternType="solid", start_color="00FF00")
blue_fill = PatternFill(patternType="solid", start_color="0000FF")
yellow_fill = PatternFill(patternType="solid", start_color="FFFF00")

##### Monkeypatch navigation properties into Cell ################


def _left(self):
    """Monkey Patch property of Cell for the cell to the left."""
    ws = self.parent
    if self.column == 1:
        raise ValueError("Already at column 1 ! Cannot move left")
    return ws.cell(row=self.row, column=self.column - 1)


setattr(opx.cell.Cell, "left", property(fget=_left))


def _right(self):
    """Monkey Patch property of Cell for the cell to the right."""
    ws = self.parent
    if self.column == ws.max_column:
        raise ValueError("Already at last column ! Cannot move right")
    return ws.cell(row=self.row, column=self.column + 1)


setattr(opx.cell.Cell, "right", property(fget=_right))


def _up(self):
    """Monkey Patch property of Cell for the cell to the left."""
    ws = self.parent
    if self.row == 1:
        raise ValueError("Already at top row ! Cannot move up")
    return ws.cell(row=self.row - 1, column=self.column)


setattr(opx.cell.Cell, "up", property(fget=_up))


def _down(self):
    """Monkey Patch property of Cell for the cell to the left."""
    ws = self.parent
    if self.row == ws.max_row:
        raise ValueError("Already at bottom row ! Cannot move down")
    return ws.cell(row=self.row + 1, column=self.column)


setattr(opx.cell.Cell, "down", property(fget=_down))


def _move(self, delta_row=0, delta_column=0):
    """Return the cell that is at the specified relative position to this cell."""
    ws = self.parent
    nr = self.row + delta_row
    nc = self.column + delta_column
    if not 1 <= nr <= ws.max_row:
        raise ValueError("Attempted to navigate to a row that is off the worksaheet")
    if not 1 <= nc <= ws.max_column:
        raise ValueError("Attempted to navigate to a column that is off the worksheet.")
    return ws.cell(row=nr, column=nc)


setattr(opx.cell.Cell, "move", _move)
####################################################################


def save_virtual_workbook(wb):
    """Get workbook to string.

    Args:
        wb (openpyxl.Workbook): Spreadsheet to save

    Returns:
        (bytes):
            Byte string representeation of xlsx file.

    Notes:
        This is a replacement for the deprecated meothd from openpyxl.writer.excel of the same name.
    """
    with NamedTemporaryFile() as tmp:
        tmp.close()  # with statement opened tmp, close it so wb.save can open it
        wb.save(tmp.name)
        with open(tmp.name, "rb") as f:
            f.seek(0)  # probably not needed anymore
            new_file_object = f.read()
    return new_file_object


def sub_value(cell, module):
    """Use a cell value of te form self.attr.attr|format to replace it with a number."""
    parts = str(cell.value).split("|")
    if len(parts) == 1:
        parts.append("0")
    root = module
    for attr in parts[0].split(".")[1:]:
        root = getattr(root, attr)
    if not isinstance(root, (str, float, int)):
        root = str(root)
    cell.value = root
    if isinstance(root, (int, float)):
        cell.data_type = "n"
    else:
        cell.data_type = "s"
    cell.number_format = parts[1]


class BaseSpreadsheet:
    """Base class for manipulating spreadsheets."""

    def __init__(self, filename, blank=False):
        """Create the spreadsheet object or load from disk."""
        self.workbook = opx.load_workbook(filename, data_only=not blank)
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        self.errors = []

    @abstractmethod
    def get_name(self):
        """Return a string representing the name of trhis spreadsheet."""
        raise NotImplementedError("Must provide a get_name method")

    def search(self, *text, start_row=1, start_col=1, from_cell=None):
        """Search for text in the current worksheet.

        Args:
            text  (string): Cell contents to locate.

        Keyword Arguments:
            start_row (int):
                Row to start search from (see also from_cell)
            start_col (int):
                Column to start search from (see also from_cell)
            from_cell (Cell):
                Start the search from this cell, overriding start_row and start_col

        Returns:
            (cell) the openpyxl cell containing the text.
        """
        if isinstance(from_cell, opx.cell.Cell):  # Override start_row and start_col
            start_row = from_cell.row
            start_col = from_cell.column
        for ix in range(start_row, self.sheet.max_row + 1):  # Loop over rows
            for iy in range(start_col, self.sheet.max_column + 1):  # Loop over columns
                cell = self.sheet.cell(row=ix, column=iy)
                if str(cell.value).strip() == text[0].strip():  # First test matches
                    for iz, text in enumerate(text[1:]):  # Loop over next tests
                        cell2 = self.sheet.cell(row=ix + iz + 1, column=iy)  # look at rows down
                        if str(cell2.value).strip() != text[0].strip():  # If test doesn't match, break out of tests
                            break
                    else:  # Didn't break out of tests, so return cell found as all tests match
                        return cell
                    if len(text) != 1:  # If multiple tests then one of them failed, so carry on looking
                        continue
                    return cell  # One test and it amtched so return cell
        return None

    def is_non_blank(self, cell):
        """Return True if cell is not blank."""
        cell = cell.value
        try:
            cell = str(cell).strip()
        except (TypeError, ValueError):
            raise RuntimeError("Oops, found a cell that didn't want to be a string!")
        return cell != ""

    def respond(self):
        """Create a stream for the current workbook."""
        response = HttpResponse(
            content=save_virtual_workbook(self.workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename={self.get_name()}"
        return response

    def as_file(self, dirname):
        """Save the workbook to a particular path."""
        if not os.path.exists(dirname):
            os.makedirs(dirname, exist_ok=True)
        filename = os.path.join(dirname, self.get_name())
        self.workbook.save(filename)
        return filename

    @abstractmethod
    def fill_in(self, *args, **kwargs):
        """Fill in a spreadsheet from some data sources."""
        raise NotImplementedError("fil_in needs to be implemented in concrete subclasses.")


class Spreadsheet(BaseSpreadsheet):
    """Class to handle working with Module Marksheets using openpyxl."""

    def __init__(self, filename, blank=False):
        """Create the spreadsheet object and initialise if as a marksheet."""
        super().__init__(filename, blank)
        if not blank:
            self.sids = self.find_sid_cells()
        else:
            self.sids = None
        self.mod = f"{config.SUBJECT_PREFIX}LXXX"
        self.components = self.get_components()

    @property
    def first_row(self):
        """Return the first row which has student IDs."""
        sidcell = self.search(str(list(self.sids.keys())[0]))
        if sidcell is None:
            raise RuntimeError("Can't find the first student.")
        return sidcell.row

    @property
    def module_code(self):
        """Return the module code for this marksheet."""
        try:
            ret = str(self.search("Module Code:").right.value).replace(" ", "")
            if ret.endswith(")"):
                return ret[:-4]
            return ret
        except ValueError:
            raise RuntimeError("Uploaded Workbook is missing a module code!")

    @property
    def module_exam_code(self):
        """Return the module exam code."""
        try:
            return int(self.search("Module Exam Code:").right.value).replace(" ", "")
        except ValueError:
            pass
        try:
            ret = str(self.search("Module Code:").right.value).replace(" ", "")
            if ret.endswith(")"):
                return int(ret[-3:-1])
            else:
                return 1
        except ValueError:
            raise RuntimeError("Uploaded Workbook is missing a module exam code!")

    @property
    def module_title(self):
        """Return the module title."""
        try:
            return str(self.search("Module Title:").right.value).strip()
        except ValueError:
            raise RuntimeError("Uploaded Workbook is missing a module title!")

    @property
    def student_names(self):
        """Return the student names."""
        cell = self.search("Name")
        if cell is None:
            raise RuntimeError("Unable to locate the names of the students - no cell containing 'Name'?")
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def student_numbers(self):
        """Return the student numbers."""
        if self.sids is None:
            self.sids = self.find_sid_cells()
        return self.sids

    @property
    def student_programmes(self):
        """Return the student programmes."""
        cell = self.search("Programme")
        if cell is None:
            raise RuntimeError("Unable to locate the Programmes of the students - no cell 'Programme'")
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def student_status(self):
        """Return the student status records."""
        cell = self.search("Status")
        if cell is None:
            raise RuntimeError("Unable to locate the Status fields of the students - no cell 'Status'?")
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def student_marks(self):
        """Return the student total marks."""
        cell = self.search("Total", "%")
        if cell is None:
            raise RuntimeError("Unable to locate the total score fields of the students- no cell 'Total'?")
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def student_mark_codes(self):
        """Return an iterator of tuples of (student mark,code).

        mark will either be a float or the string "AB".
        code will either be "", "v-code" or "c-code".

        Handles both new and old spreadsheets.
        """
        cell = self.search("Code")
        codes = []
        marks = []

        if cell is not None:
            c = cell.col_idx
            ix = self.first_row
            for iy in range(ix, ix + len(self.sids)):
                codes.append(str(self.sheet.cell(column=c, row=iy).value).upper())
                if codes[-1] == "NONE":
                    codes[-1] = ""
            for ix, mk in enumerate(self.student_marks):
                try:
                    marks.append(float(mk))
                except ValueError:
                    if mk is None or str(mk) == "None":
                        marks.append(np.NaN)
                    elif str(mk).strip().startswith("AB"):
                        marks.append("AB")
                        codes[ix] += "AB"
                    else:
                        raise RuntimeError(
                            "New Style nmarksheets should only have numbers of AB in the Total column! {} ({})".format(
                                mk, type(mk)
                            )
                        )
            return list(zip(marks, codes))
        else:
            for mk, cmnt in zip(self.student_marks, self.student_comments):
                try:
                    marks.append(float(mk))
                    if cmnt.strip().lower() == "v":
                        codes.append("V")
                    elif cmnt.strip().lower() == "c":
                        codes.append("C")
                    else:
                        codes.append("")
                except ValueError:
                    if str(mk).strip().upper().startswith("AB"):
                        marks.append("AB")
                        codes.append("")
                    elif re.compile(r"\d+v").match(str(mk).lower().strip()) or cmnt.strip().lower() == "v":
                        codes.append("V")
                        res = re.compile(r"(\d+)v?").match(str(mk).lower().strip())
                        marks.append(float(res.group(1)))
                    elif re.compile(r"\d+c").match(str(mk).lower().strip()) or cmnt.strip().lower() == "v":
                        codes.append("C")
                        res = re.compile(r"(\d+)v?").match(str(mk).lower().strip())
                        marks.append(float(res.group(1)))
                    else:
                        raise RuntimeError(f"Unable to interpret mark and code {mk} {cmnt}")
            return zip(marks, codes)

    @property
    def student_comments(self):
        """Return the comments."""
        for label in ["Moderated", "Comment"]:
            cell = self.search(label)
            if cell is not None:
                break
        else:
            raise RuntimeError("Unable to locate the comments column of the students - no cell 'Moderated'?")
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def coursework_marks(self):
        """Return the coursework marks - where there is a component called Coursework."""
        try:
            cell = self.search("Course")
        except (ValueError, TypeError):
            return np.ones(len(self.sids)) * np.NaN
        if cell is None:
            return np.ones(len(self.sids)) * np.NaN
        names = []
        c = cell.col_idx
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            names.append(str(self.sheet.cell(column=c, row=iy).value))
        return names

    @property
    def section_A_marks(self):
        """Find the marks in a Section A."""
        A_total = 0.4
        try:
            cell = self.search("Section/Question:  ")
            col_idx = []
            A_total = 0.0
            for col in range(cell.col_idx, cell.col_idx + 15):
                if str(self.sheet.cell(column=col, row=cell.row).value).upper().strip() == "A":
                    col_idx.append(col)
                    try:
                        A_total += float(self.sheet.cell(column=col, row=cell.row + 2).value) / 100.0
                    except (ValueError, TypeError, AttributeError):
                        pass
        except (ValueError, TypeError):
            return np.ones(len(self.sids)) * np.NaN
        if cell is None or not len(col_idx):
            return np.ones(len(self.sids)) * np.NaN
        names = []
        ix = self.first_row
        for iy in range(ix, ix + len(self.sids)):
            A_mark = 0
            for col in col_idx:
                v = self.sheet.cell(column=col, row=iy).value
                if v is not None:
                    try:
                        A_mark += float(v) / A_total
                    except (ValueError, TypeError, ZeroDivisionError):
                        pass
                names.append(str(A_mark))
        return names

    @property
    def entries(self):
        """Return all the entries for this marksheet."""
        for record in zip(
            self.student_names,
            self.student_numbers,
            self.student_programmes,
            self.student_status,
            self.student_mark_codes,
            self.coursework_marks,
            self.section_A_marks,
            self.student_comments,
        ):
            yield {
                k: v
                for k, v in zip(
                    ["name", "number", "programme", "status", "mark", "coursework", "section_A", "comment"], record
                )
            }

    def get_name(self):
        """Module name is the spreadsheet name."""
        return f"{self.mod}.xlsx"

    def is_sid(self, cell):
        """Return true if cell contents looks like a student ID."""
        cell = cell.value
        try:
            cell = int(cell)
        except (TypeError, ValueError):
            return False
        return 200000000 < cell < 210000000

    def is_mark(self, cell):
        """Return True for cells that contain numbers between 0 and 100."""
        cell = cell.value
        try:
            cell = float(cell)
        except (TypeError, ValueError):
            return False
        return 0 <= cell <= 100

    def get_components(self):
        """Return a dictionary of component:cell."""
        ret = dict()
        cell = self.search("Section/Question:  ")
        start_idx = cell.col_idx + 2
        row = cell.row + 1
        for col in range(start_idx, start_idx + 15):
            name = str(self.sheet.cell(column=col, row=row).value).strip()
            value = self.sheet.cell(column=col, row=row)
            if name.upper() == "NONE":
                break
            ret[name] = value
        return ret

    def find_sid_cells(self):
        """Return a list of cells containing student IDs."""
        sidcell = self.search("SID")
        if sidcell is None:
            raise RuntimeError(
                "Uploaded Spreadsheet was not a module marksheet - could not find the Student ID Column."
            )
        c = sidcell.col_idx
        for ix in range(sidcell.row, self.sheet.max_row + 1):
            if self.is_sid(self.sheet.cell(row=ix, column=c)):
                break
        else:
            raise RuntimeError("Fell off the end of the spreadsheet before finding any student IDs!")
        sids = {}
        for ix in range(ix, self.sheet.max_row + 1):
            if self.is_sid(self.sheet.cell(row=ix, column=c)):
                cell = self.sheet.cell(row=ix, column=c)
                sids[int(cell.value)] = cell
            else:
                break
        return sids

    def get_comp_mark(self, component, student):
        """Return the mark in the spreadsheet for the component and student."""
        sid = student.number
        comp = component.header
        try:
            row = self.sids[sid].row
        except IndexError:
            raise ValueError(f"Cannot ind a student {student} in the spreadsheet!")
        try:
            col = self.components[comp].col_idx
        except IndexError:
            raise ValueError(
                f"Cannot find a column for component {component} - the spreadsheet must"
                + " match the website's list of module components."
            )
        return self.sheet.cell(row=row, column=col).value

    def fill_in(self, *args, **kwargs):
        """Fill in details from a module."""
        module = args[0]
        entries = kwargs.get("entries", None)

        # Iterate over all worksheets in the template
        for sheet in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet]

            # fill in all cells from module attributes
            for ix in range(1, self.sheet.max_row + 1):  # Loop over rows
                for iy in range(1, self.sheet.max_column + 1):  # Loop over columns
                    cell = self.sheet.cell(row=ix, column=iy)
                    if str(cell.value).startswith("self."):
                        sub_value(cell, module)

            # Now fill in components
            mh = 0
            component_columns = {}
            for c, ix in zip(module.VITALS.all(), range(5, 5 + module.VITALS.count())):
                component_columns[c] = (ix, c.name)
                cell = self.sheet.cell(row=11, column=ix)
                cell.value = c.name
                mh = max(len(c.name), mh)
                if len(c.name) > 2:
                    alignment = cell.alignment.copy(text_rotation=90)
                    cell.alignment = alignment
                else:
                    alignment = cell.alignment.copy(text_rotation=0)
                    cell.alignment = alignment
                cell = self.sheet.cell(row=12, column=ix)
                cell.value = "P/F"
                cell = self.sheet.cell(row=13, column=ix)
                cell.value = 1.0
            if mh > 2:
                self.sheet.row_dimensions[11].height = 6.5 * mh

            # Now fill in students
            cell = self.search("V/C")
            if entries is None:
                entries = module.student_enrollments.prefetch_related("student", "student__vital_results").order_by(
                    "student"
                )
            elif isinstance(entries, dict):
                entries = (
                    module.student_enrollments.filter(**entries)
                    .prefetch_related("student", "student__vital_results")
                    .order_by("student")
                )
            else:
                entries = (
                    module.student_enrollments.filter(entries)
                    .prefetch_related("student", "student__vital_results")
                    .order_by("student")
                )

            for ix, ent in enumerate(entries):
                row = ix + 16
                self.sheet.cell(row=row, column=1).value = ent.student.display_name
                self.sheet.cell(row=row, column=2).value = ent.student.programme.name
                self.sheet.cell(row=row, column=3).value = ent.student.number
                self.sheet.cell(row=row, column=4).value = ent.status.code
                for _, (comp_col, mtype) in component_columns.items():
                    mks = ent.student.vital_results.filter(vital__name=mtype)
                    comp_mark = mks.count() > 0 and mks.last().passed
                    self.sheet.cell(row=row, column=comp_col).value = "P" if comp_mark else ""

            # Now remove excess rows
            if ix + 17 < 275:
                self.sheet.delete_rows(ix + 17, 258 - ix)
                last_row = 23 + ix
                # rewrite formulae
                for r in range(last_row + 1, last_row + 6):
                    for c in range(5, 20):
                        cell = self.sheet.cell(column=c, row=r)
                        if isinstance(cell.value, str):
                            cell.value = cell.value.replace("281", str(last_row))
                            cell.value = cell.value.replace("285", str(last_row + 4))
                cell = self.sheet.cell(column=3, row=7)
                cell.value = cell.value.replace("285", str(last_row + 4))

            self.sheet.title = module.code
            self.mod = module.code
        self.sheet = self.workbook[self.workbook.sheetnames[0]]


class TutorReportSheet(BaseSpreadsheet):
    """Spreadsheet for the Tutor report."""

    def __init__(self, filename, blank=False, tutor=None, tests_class=None, vitals_class=None):
        """Create the spreadsheet object or load from disk."""
        super().__init__(filename, blank)
        self.vitals_sheet = self.workbook[self.workbook.sheetnames[1]]
        self.tests_sheet = self.workbook[self.workbook.sheetnames[0]]
        self.tutor = tutor
        self.Test = tests_class
        self.VITAL = vitals_class
        self.sidmap = []
        self.namemap = []
        if self.tutor:
            if self.Test is None:
                self.Test = self.tutor.tests.model
            if self.VITAL is None:
                self.VITAL = self.tutor.VITALS.model
            self.enter_student_details()
            self.enter_test_names()
            self.enter_vtials_names()
            self.enter_test_results()
            self.enter_test_attempts()
            self.enter_vital_results()

    def get_name(self):
        """Use Tutor initials as name."""
        return f"{self.tutor.initials}.xlsx"

    def _fill_in_details(self, working_cell):
        """Complete student name, programme and SID entries."""
        sidmap = {}
        for student in self.tutor.tutees.all().order_by("last_name", "first_name"):
            working_cell.value = f"{student.last_name}, {student.first_name}"
            working_cell.right.value = f"{student.programme.name}"
            working_cell.right.right.value = f"{student.number}"
            sidmap[student.number] = working_cell.row
            working_cell = working_cell.down
        self.sidmap.append(sidmap)

    def enter_student_details(self):
        """Use tutor to build a list of student details and enter into spreadsheet."""
        self.sheet = self.tests_sheet
        working_cell = self.search("Student Name").down
        self._fill_in_details(working_cell)
        working_cell = self.search("Student Name", from_cell=working_cell).down
        self._fill_in_details(working_cell)
        self.sheet = self.vitals_sheet
        working_cell = self.search("Student Name").down
        self._fill_in_details(working_cell)

    def _enter_names(self, cell, names, counts):
        """Fill in a bunch of test/vital names with merged modules."""
        start = cell
        namemap = {}
        for thing in names:
            cell.value = f"{shorten(thing.name,30)}"
            alignment = cell.alignment.copy(text_rotation=90)
            cell.alignment = alignment
            namemap[thing.pk] = cell.column
            cell = cell.right
        cell = start.up
        self.namemap.append(namemap)
        for module, number in counts.items():
            self.sheet.merge_cells(
                start_row=cell.row, start_column=cell.column, end_row=cell.row, end_column=cell.column + number - 1
            )
            cell.value = module
            cell = cell.move(delta_column=number)

    def enter_test_names(self):
        """Enter all of the test names."""
        if self.Test is None:
            return
        tests = self.Test.objects.filter(results__user__apt=self.tutor).distinct().order_by("module", "name")
        counts = {}
        for test in tests:
            counts[test.module.code] = tests.filter(module=test.module).count()
        self.sheet = self.tests_sheet
        cell = self.search("SID").right
        self._enter_names(cell, tests, counts)
        cell = self.search("SID", from_cell=cell.left.down).right
        self._enter_names(cell, tests, counts)

    def enter_vtials_names(self):
        """Enter the names of the VITALs into the second sheet."""
        if self.VITAL is None:
            return
        vitals = self.VITAL.objects.filter(student_results__user__apt=self.tutor).distinct().order_by("module", "name")
        counts = {}
        for vital in vitals:
            counts[vital.module.code] = vitals.filter(module=vital.module).count()
        self.sheet = self.vitals_sheet
        cell = self.search("SID").right
        self._enter_names(cell, vitals, counts)

    def enter_test_results(self):
        """Enter the test scores into the spreadsheet."""
        self.sheet = self.tests_sheet
        for sid, row in self.sidmap[0].items():
            student = self.tutor.tutees.get(number=sid)
            for test_score in student.test_results.all():
                column = self.namemap[0][test_score.test.pk]
                cell = self.sheet.cell(row=row, column=column)
                cell.value = test_score.score
                cell.fill = green_fill if test_score.passed else red_fill

    def enter_test_attempts(self):
        """Enter the test scores into the spreadsheet."""
        self.sheet = self.tests_sheet
        for sid, row in self.sidmap[1].items():
            student = self.tutor.tutees.get(number=sid)
            for test_score in student.test_results.all():
                attempts = test_score.attempts.count()
                column = self.namemap[1][test_score.test.pk]
                cell = self.sheet.cell(row=row, column=column)
                cell.value = attempts
                if test_score.passed:
                    cell.fill = green_fill if attempts < 2 else yellow_fill
                else:
                    cell.fill = blue_fill if attempts < 2 else red_fill

    def enter_vital_results(self):
        """Enter the test scores into the spreadsheet."""
        self.sheet = self.vitals_sheet
        for sid, row in self.sidmap[2].items():
            student = self.tutor.tutees.get(number=sid)
            for vital_result in student.vital_results.all():
                column = self.namemap[2][vital_result.vital.pk]
                cell = self.sheet.cell(row=row, column=column)
                cell.value = "P" if vital_result.passed else "F"
                cell.fill = green_fill if vital_result.passed else red_fill
